# totalasia_matcher.py
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook


# =========================
# Config
# =========================

@dataclass
class MatchConfig:
    """
    Matching configuration.

    min_common_words_primary:
        Primary threshold for common English tokens in description matching (>=3).
    min_common_words_relaxed:
        Relaxed threshold used ONLY in fallback rule 3 (>=2).
    """
    min_common_words_primary: int = 3
    min_common_words_relaxed: int = 2
    product_pad_width: int = 6   # 新增：Product 统一补齐位数（默认6）


# =========================
# Text and parsing utilities
# =========================

def english_only(text: object) -> str:
    """
    Keep English/number tokens only:
    - lowercase
    - remove punctuation and non [a-z0-9 whitespace]
    - ignore Chinese automatically because it is removed
    """
    if pd.isna(text):
        return ""
    s = str(text).lower()
    # normalize curly apostrophe
    s = s.replace("’", "'")
    # keep only a-z,0-9, whitespace
    s = re.sub(r"[^a-z0-9\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def tokenize_en(text: object) -> List[str]:
    """Tokenize cleaned English text into list of tokens."""
    s = english_only(text)
    if not s:
        return []
    return [t for t in s.split(" ") if t]


def common_token_count(a_tokens: List[str], b_tokens: List[str]) -> int:
    """Count number of common tokens between two token lists (set intersection)."""
    return len(set(a_tokens) & set(b_tokens))


def desc_similarity(a_en: str, b_en: str) -> float:
    """
    Similarity score for ranking candidates.
    token_set_ratio is robust to:
    - different word order
    - missing/extra words
    """
    return float(fuzz.token_set_ratio(a_en, b_en))


def extract_digits(x: object) -> Optional[int]:
    """
    Extract first continuous digit sequence as int.
    Returns None if no digits found.
    """
    if pd.isna(x):
        return None
    m = re.search(r"\d+", str(x))
    return int(m.group()) if m else None


_UNIT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(kg|g|ml|l|pcs)\b", re.IGNORECASE)


def parse_pack_size(pack: object) -> str:
    """
    Parse Supplier 'PACK SIZE' to single-unit size string.

    Rules:
    - If contains 'x' (or ×), take the segment after the last x:
        '2x5L' -> '5l'
        '12x540g bag' -> '540g'
    - Else take first number+unit:
        '20kg bag' -> '20kg'
    """
    if pd.isna(pack):
        return ""
    s = str(pack).strip().lower().replace("×", "x")

    if "x" in s:
        tail = s.split("x")[-1]
        m = _UNIT_RE.search(tail)
        if m:
            return f"{m.group(1)}{m.group(2)}".lower()

    m = _UNIT_RE.search(s)
    if m:
        return f"{m.group(1)}{m.group(2)}".lower()

    return ""


def norm_size(s: object) -> str:
    """Normalize size string for comparison."""
    if pd.isna(s):
        return ""
    x = str(s).strip().lower().replace(" ", "")
    x = x.replace("litre", "l").replace("liter", "l")
    return x


def product_key_raw(x: object) -> Optional[str]:
    """
    Keep ERP Product ID as a string "as-is".
    - preserves existing leading zeros
    - DOES NOT pad to a unified length (avoid extra zeros)
    """
    if pd.isna(x):
        return None
    s = str(x).strip().replace(",", "")
    if s.endswith(".0"):
        s = s[:-2]
    # scientific notation safeguard
    if "e" in s.lower():
        try:
            s = str(int(float(s)))
        except Exception:
            pass
    return s


# =========================
# Loading data
# =========================

def load_supplier(path: str | Path) -> pd.DataFrame:
    """
    Load Supplier Excel and standardize columns:
    CODE, DESCRIPTION, PACK SIZE, QTY, PRICE
    """
    df = pd.read_excel(path)
    df.columns = [c.strip() for c in df.columns]

    rename = {
        "CODE": "code",
        "DESCRIPTION": "desc_a",
        "PACK SIZE": "pack_size_a",
        "QTY": "qty_a",
        "PRICE": "price_a",
    }
    df = df.rename(columns=rename)

    required = ["code", "desc_a", "pack_size_a", "qty_a", "price_a"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Supplier missing columns: {missing}. Found: {df.columns.tolist()}")

    df["code"] = df["code"].apply(extract_digits).astype("Int64")
    df["desc_a"] = df["desc_a"].astype(str).str.strip()
    df["pack_size_a"] = df["pack_size_a"].astype(str).str.strip()

    df["qty_a"] = pd.to_numeric(df["qty_a"], errors="coerce").astype("Int64")
    df["price_a"] = df["price_a"].astype(str).str.replace("£", "", regex=False).str.strip()
    df["price_a"] = pd.to_numeric(df["price_a"], errors="coerce")

    return df


from openpyxl import load_workbook
import pandas as pd
import re

def _product_from_excel_cell(cell) -> str:
    """
    Return Product string as it appears in Excel:
    - If cell is numeric and has a number_format like '000000'/'0000000', pad with that many zeros.
    - Otherwise return string value as-is.
    """
    v = cell.value
    if v is None:
        return ""

    # If already text, keep it
    if isinstance(v, str):
        return v.strip()

    # If numeric, check number_format for zero-padding
    if isinstance(v, (int, float)):
        # int-like float (401.0)
        iv = int(v) if float(v).is_integer() else v
        fmt = (cell.number_format or "").strip()

        # Common cases: '000000', '0000000'
        if re.fullmatch(r"0+", fmt):
            width = len(fmt)
            return str(int(iv)).zfill(width)

        # If format contains a block of zeros (less strict), take the longest consecutive zeros
        m = re.findall(r"0{2,}", fmt)
        if m:
            width = max(len(x) for x in m)
            return str(int(iv)).zfill(width)

        # No padding format -> just str(int)
        return str(int(iv)) if float(v).is_integer() else str(v)

    return str(v).strip()


def load_erp(path: str | Path, cfg: MatchConfig) -> pd.DataFrame:
    df = pd.read_excel(path)
    df.columns = [c.strip() for c in df.columns]

    rename = {
        "Product": "product",
        "Suppl category": "suppl_category",
        "Suppl Category": "suppl_category",
        "Suppl": "suppl_category",
        "Description": "desc_b",
        "Size": "size_b",
        "Pack": "pack_b",
        "Store": "store",
    }
    df = df.rename(columns=rename)

    required = ["product", "suppl_category", "desc_b", "size_b", "pack_b", "store"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"ERP missing columns: {missing}. Found: {df.columns.tolist()}")

    # supplier category as number
    def extract_digits(x):
        if pd.isna(x): return None
        m = re.search(r"\d+", str(x))
        return int(m.group()) if m else None

    df["suppl_category"] = df["suppl_category"].apply(extract_digits).astype("Int64")

    # Product: force string, remove trailing .0, then zfill to cfg.product_pad_width
    def product_key_raw(x):
        if pd.isna(x): return ""
        s = str(x).strip().replace(",", "")
        if s.endswith(".0"): s = s[:-2]
        if "e" in s.lower():
            try:
                s = str(int(float(s)))
            except Exception:
                pass
        return s

    df["product"] = df["product"].apply(product_key_raw)
    df["product"] = df["product"].apply(lambda s: (s or "").zfill(int(cfg.product_pad_width)))

    df["desc_b"] = df["desc_b"].astype(str).str.strip()
    df["size_b"] = df["size_b"].astype(str).str.strip()

    return df


def enrich_for_matching(supplier: pd.DataFrame, erp: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    supplier = supplier.copy()
    erp = erp.copy()

    supplier["desc_a_en"] = supplier["desc_a"].apply(english_only)
    supplier["tok_a"] = supplier["desc_a"].apply(tokenize_en)
    supplier["pack_parsed"] = supplier["pack_size_a"].apply(parse_pack_size)
    supplier["pack_norm"] = supplier["pack_parsed"].apply(norm_size)

    erp["desc_b_en"] = erp["desc_b"].apply(english_only)
    erp["tok_b"] = erp["desc_b"].apply(tokenize_en)
    erp["size_norm"] = erp["size_b"].apply(norm_size)

    # product is already padded string -> use directly
    erp["product_key"] = erp["product"].astype(str)

    return supplier, erp


def run_stock_check(
    supplier_path: str | Path,
    erp_path: str | Path,
    output_dir: str | Path,
    cfg: Optional[MatchConfig] = None,
    export_unmatched: bool = True,
    fill_missing_qty_zero: bool = False
):
    cfg = cfg or MatchConfig()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    supplier = load_supplier(supplier_path)
    erp = load_erp(erp_path, cfg)          # <-- cfg passed
    supplier, erp = enrich_for_matching(supplier, erp)

    matched_map, unmatched = match_supplier_to_erp(supplier, erp, cfg)
    stock_check = build_stock_check(erp, matched_map, fill_missing_qty_zero=fill_missing_qty_zero)

    stock_path = output_dir / "Stock Check.xlsx"
    with pd.ExcelWriter(stock_path, engine="openpyxl") as w:
        stock_check.to_excel(w, index=False, sheet_name="Stock Check")

    force_excel_text_column(stock_path, sheet_name="Stock Check", col_letter="B")

    if export_unmatched:
        unmatched_path = output_dir / "Unmatched.xlsx"
        with pd.ExcelWriter(unmatched_path, engine="openpyxl") as w:
            unmatched.to_excel(w, index=False, sheet_name="Unmatched")

    map_path = output_dir / "Matched_Map.xlsx"
    with pd.ExcelWriter(map_path, engine="openpyxl") as w:
        matched_map.to_excel(w, index=False, sheet_name="Matched_Map")

    return stock_check, matched_map, unmatched


# =========================
# Matching core
# =========================

def _rank_candidates(
    a_tok: List[str],
    a_en: str,
    cand_df: pd.DataFrame,
    min_common: int
) -> pd.DataFrame:
    """
    Filter and rank candidate ERP rows by:
    1) common_words desc
    2) similarity desc
    """
    commons = cand_df["tok_b"].apply(lambda bt: common_token_count(a_tok, bt))
    cands = cand_df.loc[commons >= min_common].copy()
    if cands.empty:
        return cands

    cands["common_words"] = commons[commons >= min_common].values
    cands["sim"] = cands["desc_b_en"].apply(lambda bt: desc_similarity(a_en, bt))
    cands = cands.sort_values(["common_words", "sim"], ascending=[False, False])
    return cands


def match_supplier_to_erp(
    supplier: pd.DataFrame,
    erp: pd.DataFrame,
    cfg: MatchConfig
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Apply the matching rules:

    Rule 1: (1+2)
      1) CODE matches Suppl category AND
      2) desc common words >= 3  -> confirm

    Rule 2: (2+3)
      2) desc common words >= 3 AND
      3) parsed size matches ERP size -> confirm

    Rule 3: relaxed desc, must satisfy (1+2+3) simultaneously
      1) CODE matches AND
      2) desc common words >= 2 AND
      3) size matches -> confirm

    Returns:
      matched_map: Supplier rows with confirmed ERP match
      unmatched: Supplier rows not confirmed
    """
    erp_by_code = erp.dropna(subset=["suppl_category"]).groupby("suppl_category")

    matched_rows: List[Dict] = []
    unmatched_rows: List[Dict] = []

    for i, a in supplier.iterrows():
        a_code = a["code"]
        a_tok = a["tok_a"]
        a_en = a["desc_a_en"]
        a_pack = a["pack_norm"]

        chosen = None
        rule_used = ""
        match_code = False
        match_desc = False
        match_size = False
        common_words_val = None
        sim_val = None

        # -------- Rule 1: 1+2 --------
        if pd.notna(a_code) and a_code in erp_by_code.groups:
            match_code = True
            code_cands = erp.loc[erp_by_code.groups[a_code]].copy()
            ranked = _rank_candidates(a_tok, a_en, code_cands, cfg.min_common_words_primary)
            if not ranked.empty:
                chosen = ranked.iloc[0]
                match_desc = True
                rule_used = "1:code+desc(>=3)"
                common_words_val = int(chosen["common_words"])
                sim_val = float(chosen["sim"])

        # -------- Rule 2: 2+3 --------
        if chosen is None:
            ranked = _rank_candidates(a_tok, a_en, erp, cfg.min_common_words_primary)
            if not ranked.empty:
                best = ranked.iloc[0]
                match_desc = True
                size_ok = (a_pack != "" and a_pack == norm_size(best["size_b"]))
                if size_ok:
                    chosen = best
                    match_size = True
                    rule_used = "2:desc(>=3)+size"
                    common_words_val = int(chosen["common_words"])
                    sim_val = float(chosen["sim"])

        # -------- Rule 3: relaxed, require 1+2+3 all --------
        if chosen is None:
            if pd.notna(a_code) and a_code in erp_by_code.groups:
                match_code = True
                code_cands = erp.loc[erp_by_code.groups[a_code]].copy()
                ranked = _rank_candidates(a_tok, a_en, code_cands, cfg.min_common_words_relaxed)
                if not ranked.empty:
                    best = ranked.iloc[0]
                    size_ok = (a_pack != "" and a_pack == norm_size(best["size_b"]))
                    if size_ok:
                        chosen = best
                        match_desc = True
                        match_size = True
                        rule_used = "3:code+desc(>=2)+size"
                        common_words_val = int(chosen["common_words"])
                        sim_val = float(chosen["sim"])

        confirmed = chosen is not None

        base = {
            "supplier_row": i,
            "A_CODE": a_code,
            "A_DESC": a["desc_a"],
            "A_PACK_SIZE": a["pack_size_a"],
            "A_PACK_PARSED": a["pack_parsed"],
            "A_QTY": a["qty_a"],
            "A_PRICE": a["price_a"],
            "confirmed": confirmed,
            "rule_used": rule_used,
            "match_code": match_code,
            "match_desc": match_desc,
            "match_size": match_size,
            "common_words": common_words_val,
            "sim": sim_val,
        }

        if confirmed:
            base.update({
                "ERP_Product": chosen["product_key"],   # keep leading zeros
                "ERP_Suppl_category": chosen["suppl_category"],
                "ERP_Description": chosen["desc_b"],
                "ERP_Size": chosen["size_b"],
            })
            matched_rows.append(base)
        else:
            unmatched_rows.append(base)

    return pd.DataFrame(matched_rows), pd.DataFrame(unmatched_rows)


# =========================
# Excel post-processing
# =========================

def force_excel_text_column(xlsx_path: str | Path, sheet_name: str, col_letter: str) -> None:
    """
    Force an Excel column to TEXT number format ('@') and stringify values.
    This prevents Excel from dropping leading zeros when opening the file.
    """
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    # Skip header row
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{col_letter}{row}"]
        if cell.value is None:
            continue
        cell.number_format = "@"
        cell.value = str(cell.value)

    # Also set entire column format
    for cell in ws[col_letter]:
        cell.number_format = "@"

    wb.save(xlsx_path)


# =========================
# Stock Check build/export
# =========================

def build_stock_check(
    erp: pd.DataFrame,
    matched_map: pd.DataFrame,
    fill_missing_qty_zero: bool = False
) -> pd.DataFrame:
    """
    Build Stock Check output:
    - Row order EXACTLY equals ERP order
    - Columns:
        Index, Product, Suppl category, Description, Size, Pack, Store, Price, Qty
    - Price/Qty come from Supplier.
    - Optional: fill missing Qty with 0 (supplier out of stock).
    """
    m = matched_map.dropna(subset=["ERP_Product"]).copy()
    # If duplicates occur, keep the first mapping
    m = m.drop_duplicates(subset=["ERP_Product"], keep="first")

    qty_map = dict(zip(m["ERP_Product"], m["A_QTY"]))
    price_map = dict(zip(m["ERP_Product"], m["A_PRICE"]))

    out = erp.copy()
    if "product_key" not in out.columns:
        out["product_key"] = out["product"].apply(product_key_raw)

    out["Qty"] = out["product_key"].map(qty_map)
    out["Price"] = out["product_key"].map(price_map)

    if fill_missing_qty_zero:
        out["Qty"] = out["Qty"].fillna(0).astype(int)

    stock_check = pd.DataFrame({
        "Product": out["product_key"],
        "Suppl category": out["suppl_category"],
        "Description": out["desc_b"],
        "Size": out["size_b"],
        "Pack": out["pack_b"],
        "Store": out["store"],
        "Price": out["Price"],
        "Qty": out["Qty"],
    })

    # First column as Index
    stock_check.insert(0, "Index", range(1, len(stock_check) + 1))

    # Ensure Product is string in dataframe (Excel still may coerce; we post-fix with openpyxl)
    stock_check["Product"] = stock_check["Product"].astype(str)

    return stock_check


def run_stock_check(
    supplier_path: str | Path,
    erp_path: str | Path,
    output_dir: str | Path,
    cfg: Optional[MatchConfig] = None,
    export_unmatched: bool = True,
    fill_missing_qty_zero: bool = False
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    End-to-end runner:
    1) Load Supplier + ERP
    2) Enrich for matching
    3) Match
    4) Build Stock Check (ERP order)
    5) Export Stock Check (+ optional Unmatched + audit map)

    Exports in output_dir:
      - Stock Check.xlsx
      - Unmatched.xlsx (optional)
      - Matched_Map.xlsx (audit)
    """
    cfg = cfg or MatchConfig()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    supplier = load_supplier(supplier_path)
    erp = load_erp(erp_path,cfg)
    supplier, erp = enrich_for_matching(supplier, erp)

    matched_map, unmatched = match_supplier_to_erp(supplier, erp, cfg)
    stock_check = build_stock_check(erp, matched_map, fill_missing_qty_zero=fill_missing_qty_zero)

    # Export
    stock_path = output_dir / "Stock Check.xlsx"
    with pd.ExcelWriter(stock_path, engine="openpyxl") as w:
        stock_check.to_excel(w, index=False, sheet_name="Stock Check")

    # Force Product column (B) to TEXT to preserve leading zeros in Excel UI
    # Column A is Index, so Product is column B.
    force_excel_text_column(stock_path, sheet_name="Stock Check", col_letter="B")

    if export_unmatched:
        unmatched_path = output_dir / "Unmatched.xlsx"
        with pd.ExcelWriter(unmatched_path, engine="openpyxl") as w:
            unmatched.to_excel(w, index=False, sheet_name="Unmatched")

    map_path = output_dir / "Matched_Map.xlsx"
    with pd.ExcelWriter(map_path, engine="openpyxl") as w:
        matched_map.to_excel(w, index=False, sheet_name="Matched_Map")

    return stock_check, matched_map, unmatched